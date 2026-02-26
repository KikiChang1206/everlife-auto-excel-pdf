import streamlit as st
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
import io

# 定義格線樣式 (細線)
thin_side = Side(style='thin')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

def process_invoice(file_bytes):
    # 讀取 Excel，data_only=True 確保讀到公式結果
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active 

    # --- 1. 公司資訊更換 (A2-A4) ---
    company_val = str(ws['A2'].value) if ws['A2'].value else ""
    if 'EVERLIFE-AL' in company_val:
        ws['A2'] = '歐瑞生醫科技有限公司 Allre Biological Technology Co., Ltd.'
        ws['A3'] = 'TEL : (02)29531399'
        ws['A4'] = 'Adress : 新北市板橋區中山路一段69號十樓'
    elif 'EVERLIFE-MK' in company_val:
        ws['A2'] = '蜜凱生技有限公司 MK BIOTECHNOLOGY Co., Ltd'
        ws['A3'] = 'TEL : (02)29531399'
        ws['A4'] = 'Adress : 236新北市土城區永豐路96巷8號'

    # --- 2. 畫格線與格式化品項區域 (從第 12 列開始) ---
    last_row = 12
    # 先找出資料真正的最後一列 (以 A 欄品項編號為準)
    for r in range(12, ws.max_row + 1):
        if ws.cell(row=r, column=1).value or ws.cell(row=r, column=9).value:
            last_row = r

    # 針對 A12 到 I(最後一列) 畫格線並設定對齊
    for row in ws.iter_rows(min_row=12, max_row=last_row, min_col=1, max_col=9):
        for cell in row:
            cell.border = thin_border
            # 預設置中
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            
            # 特殊處理：Description (B, C 欄) 靠左
            if cell.column_letter in ['B', 'C']:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
            # 特殊處理：金額相關 (H, I 欄) 靠右
            if cell.column_letter in ['H', 'I']:
                cell.alignment = Alignment(horizontal='right', vertical='center')

    # --- 3. 自動調整欄寬 (優化長度計算) ---
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        
        # 只掃描前 20 列來決定寬度，避免後端備註太長導致欄位炸開
        for cell in col[:20]:
            try:
                if cell.value:
                    val_str = str(cell.value)
                    # 中文字元計算長度為 2，英文為 1
                    length = sum(2 if ord(char) > 127 else 1 for char in val_str)
                    if length > max_length:
                        max_length = length
            except: pass
        
        # 根據欄位特性給予寬度補償
        if column_letter in ['B', 'C']: # Description 欄位給寬一點
            ws.column_dimensions[column_letter].width = min(max_length + 5, 40)
        else:
            ws.column_dimensions[column_letter].width = max_length + 3

    # --- 4. Terms 寫入 ---
    terms_row = last_row + 2
    ws.cell(row=terms_row, column=1, value="Terms : FOB")
    ws.cell(row=terms_row, column=1).font = Font(bold=False)
    ws.cell(row=terms_row, column=1).alignment = Alignment(horizontal='left')

    # 輸出檔案
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# Streamlit 介面
st.set_page_config(page_title="報單格式優化器", layout="centered")
st.title("🚢 報單格式優化工具")
st.write("此版本強化了格線繪製、B/C 欄靠左對齊以及欄寬自動補償。")

uploaded_file = st.file_uploader("請上傳原始報單 Excel", type=["xlsx"])

if uploaded_file:
    try:
        processed_data = process_invoice(uploaded_file.read())
        st.success("✅ 格式優化完成！")
        st.download_button(
            label="📥 下載最終版報單",
            data=processed_data,
            file_name=f"Final_{uploaded_file.name}",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        st.error(f"處理失敗，錯誤訊息：{e}")
